"""
Kate Yurmanovych
Lesson 9 Task 2
2) Создать модуль, который будет заполнять базу данных
случайными валидными значениями (минимум 100 студентов).
"""

from Lesson_9.utilities import context_manager as cm
import openpyxl
from random import choice


class StudentGenerator:

    def __init__(self):
        self.names_data = openpyxl.load_workbook("names_data.xlsx")["Sheet1"]
        self.db = "students.db"
        self.sql_tutors = "SELECT id FROM tutors"
        self.sql_faculty = "SELECT id FROM faculty"
        self.sql_insert_students = "INSERT INTO students ('name', 'surname', 'group_num', 'faculty_id', 'tutor_id') " \
                                   "VALUES (?,?,?,?,?)"
        self.groups = [1, 2, 3, 4]

    def define_existent_data(self, sql):
        with cm.MyDBManager(self.db) as db:
            result = db.execute(sql).fetchall()
            id_list = [x[0] for x in result]
        return id_list

    def fill_students_table(self, num_of_records):
        names = [str(self.names_data[x][0].value) for x in range(1, self.names_data.max_row)]
        surnames = [str(self.names_data[x][1].value) for x in range(1, self.names_data.max_row)]
        faculties_id = self.define_existent_data(self.sql_faculty)
        tutors_id = self.define_existent_data(self.sql_tutors)

        with cm.MyDBManager(self.db, "w") as db:
            for record in range(num_of_records):
                db.executemany(self.sql_insert_students, [(choice(names), choice(surnames), choice(self.groups),
                                                           choice(faculties_id), choice(tutors_id))])


if __name__ == "__main__":
    StudentGenerator().fill_students_table(100)
