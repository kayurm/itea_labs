import openpyxl
from random import choice
from Lesson_10.application.models import HandleDB, Author, Tag
from random import randint


class Seeder:

    def __init__(self):
        self.names_data = openpyxl.load_workbook("names_data.xlsx")["Sheet1"]
        self.tags_posts_data = openpyxl.load_workbook("tags_posts_data.xlsx")["Sheet1"]

    def seed_authors(self, num_of_records):
        names = [str(self.names_data[x][0].value) for x in range(1, self.names_data.max_row)]
        surnames = [str(self.names_data[x][1].value) for x in range(1, self.names_data.max_row)]
        authors = list()
        for record in range(num_of_records):
            authors.append({"name": choice(names), "surname": choice(surnames)})
        HandleDB().fill_author_collection(authors_list=authors)

    def seed_tags(self, num_of_records):
        titles = [str(self.tags_posts_data[x][0].value) for x in range(1, self.tags_posts_data.max_row)]
        tags = list()
        for record in range(num_of_records):
            tags.append({"tag": choice(titles)})
        HandleDB().fill_tag_collection(tags_list=tags)

    def seed_posts(self, num_of_records):
        post_title = [str(self.tags_posts_data[x][1].value) for x in range(1, self.tags_posts_data.max_row)]
        post_body = [str(self.tags_posts_data[x][2].value) for x in range(1, self.tags_posts_data.max_row)]
        authors = Author.objects.all()
        tags = Tag.objects.all()
        posts = list()
        for record in range(num_of_records):
            tags_random_amount = randint(0, 5)
            posts.append({"title": choice(post_title),
                          "body": choice(post_body),
                          "tag": [choice(tags) for _ in range(tags_random_amount)],
                          "author": choice(authors)})
        HandleDB().fill_post_collection(posts_list=posts)


if __name__ == "__main__":
    seeder = Seeder()
    seeder.seed_authors(10)
    seeder.seed_tags(10)
    seeder.seed_posts(10)
